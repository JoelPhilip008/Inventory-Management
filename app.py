#app.py

from flask import Flask, render_template, request, redirect, jsonify, send_from_directory, session, url_for
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl, threading, os
from pathlib import Path
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'your-very-secret-key-change-me'

# --- User Definitions ---
USERS = {
    'admin': 'scrypt:32768:8:1$kGuZ08OqCJIgKr47$50e763c6250dd62d6d3529340844b6e3d1548b42b3b5324d6be64c7f5cb9df4ea834736076b61ad6d308e08cd4b727f9d9f04b5aefd3f325eeaff2864e106e3a',
    'chindu': 'scrypt:32768:8:1$6YhnuSBnOfW06VVG$5c7f9b4ef9bebd921ddaafc76018785bebb8c2043bc89c99c2c92911c228be7d756b4ff410b5036cda19c2fe7baa05ad87680e1ad851e25d0e9e4aae1d3ced5a',
    'joel': 'scrypt:32768:8:1$kvkXBSUvTQkIdGEf$d6c4a57bb38708711d603732007d9387723ae2d80c58c6ac0b8586fd59c44a2d85d9d9524a74359d500e3b471a5ff4e79bbfeffced042c9acf319da7d1f6f82f',
    'justin': 'scrypt:32768:8:1$hrPdnSYR22tkY6Ys$c0bdfbdb8fe8c5639a3e2c89f538dceefc896c24ae3631e65df83b3e5c9b27770577c763ae95056c5a164c2061db3835eb9db4ed5751a7c5826644f73646c302',
    'arjun': 'scrypt:32768:8:1$8cjoI1O89RhGLeLJ$984518f3817fa90d7d6b3e0d1b88e65c83e1e0bf8a3b760c49a6933bbd420ae925b16771dafed6e325c926e0ea147973dbeccd9096f8a87d9a5a796138dd34d6',
    'dayal': 'scrypt:32768:8:1$KsXuzxaino4eqOsQ$caafd67796c4d5566df9aacea97fca30f2759196d32f7001f4371e6aaee1b81829e08d504bc8f04c2425fb64c12c874006427208cc10c0f9e400c3c4e7827bf2',
    'vivek': 'scrypt:32768:8:1$lieQZDppBs6PCNag$83df3464f75f8621c00cd167b916a47a9e98f6db0a0b44330a1d173702df6b3ce2776ed0fc6c136a4c8223c9b5a7894ca815ecdacf4c00a3a23c373c0b393142',
    'adarsh': 'scrypt:32768:8:1$MkiIdv2Sr0rY3EYn$d54d994c5ab8321b6cac42fc4323dd26d57dc178e02d2816889f6b1b670fd8756a9c16ab7654fe8b0a6f3c3e082f155d782593bc991697930b1b86a8492304bf',
    'jedin': 'scrypt:32768:8:1$lSGpcw2n6g4ygUME$6298b71312e0fa772526e424e6dc4d210e51eb1283b16833b95dd83b01681629a3e1b9cf5f25c90130ab18dc1c5ad7a29f787b069c679e953e8360cfe9b92320',
    'abel': 'scrypt:32768:8:1$AHOEvwseEaQBvqPY$f51846420e64d6b703069c13e589d63f2262c2864797a209b19396771ae44edcef44d19c9c3dee1baf1e5323a762ed344881714266d2bfd87ec8ca75dfa4688c',
    'fahad': 'scrypt:32768:8:1$uzYQr8e4FYY4a3JF$885bf54cfaf516aabb0c63b9a5417a1366010032a3b20268395b5649a95e35b94fdd564341e12b0295e34451c82c6bf4260c5c30054ab89537c9ccd2c26a1abb',
    'athul': 'scrypt:32768:8:1$ZWqpWLl51SipkD7Q$47abf34c471482fec32ba6327d649405a5f7feef619a9a2a270bfaf66a70532fedd3ed3ea7c0b4ef464392f8b15f3d8ee258ed489204b9225ab8f60ce08effac',
    'akhil': 'scrypt:32768:8:1$c97iVwQdFbtghSyy$1d078b25b0f8f9b835ba7a27c1b956c4f54448a1287efd550e9098491a826a3f3ca5726539323ed8e92de52bcc20c4607171c5e3d6cce755ac9b5e9b6d20438d',
    'vishnu': 'scrypt:32768:8:1$S0H2PVad8W0EHKbA$7dffb779b3b88a53fc5d735ec36da1b3c5e4df6da7c3c1f22bdbbd443804421a9121506aa5bc62371989724354d18b18803609ff77e36f01ef093ee0e03f8eec',
    'user1': 'scrypt:32768:8:1$HUBQdYyx2WvkuWGC$3240bcdb1c4b72c9f4df63e0f0630a150076e6b11379197bd4f1e295bfb71c2edc0353c06f9fa795671e499fb172adb34aa85b545ba284f7f5cccdb0de0ca907',
}
# This list defines who gets the 'admin' role upon successful login.
ADMIN_USERS = ['admin', 'chindu', 'dayal']


LOCK = threading.Lock()
BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / 'inventory.xlsx'
LOG_FILE = BASE_DIR / 'transaction_logs.xlsx' # Renamed for clarity

# --- Decorators ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            # If the request is a JSON API call, return a JSON error
            if request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html:
                return jsonify({'ok': False, 'error': 'Authentication required. Please log in again.'}), 401 # Unauthorized
            # Otherwise, for normal page loads, redirect to the login page
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'admin':
            # --- THIS IS THE CORRECTED LOGIC ---
            # Check if the request expects a JSON response (like our fetch calls do)
            if request.accept_mimetypes.accept_json and \
               not request.accept_mimetypes.accept_html:
                return jsonify({'ok': False, 'error': 'Admin access required'}), 403 # Forbidden
            
            # Otherwise, for regular page loads, redirect to the login page
            return redirect(url_for('login'))
            
        return f(*args, **kwargs)
    return decorated_function

# --- Helper Functions ---
def ensure_excel():
    if not EXCEL_FILE.exists():
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])
        ws = wb.create_sheet(title='Default')
        headers = ['Item No','Section','Section S.No','Component','Box No','Specifications','Existing Units','Part number','SKU ID','Remarks','URLS']
        ws.append(headers)
        ws.append([1,'Default',1,'Sample Item','A1','A sample component','10','PN123','SKU123','In Stock','http://example.com'])
        wb.save(EXCEL_FILE)

def ensure_log_file():
    if not LOG_FILE.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Logs'
        # Corrected Headers
        headers = ['Timestamp', 'Username', 'Transaction Type', 'Item No', 'Component', 'Quantity Changed', 'New Quantity']
        ws.append(headers)
        wb.save(LOG_FILE)

def load_workbook():
    ensure_excel()
    return openpyxl.load_workbook(EXCEL_FILE)

def headers_and_rows():
    wb = load_workbook()
    if not wb.sheetnames: return [], []
    headers = [cell.value for cell in wb.worksheets[0][1]]
    all_data = [list(row) for ws in wb.worksheets for row in ws.iter_rows(min_row=2, values_only=True)]
    return headers, all_data
    
def save_workbook_with_reindex(wb):
    # This function is correct
    if not wb.sheetnames: wb.save(EXCEL_FILE); return
    headers = [cell.value for cell in wb.worksheets[0][1]]
    item_no_col_idx = headers.index('Item No')
    section_col_idx = headers.index('Section')
    section_s_no_col_idx = headers.index('Section S.No')
    global_item_no_counter = 1
    for ws in wb.worksheets:
        section_s_no_counter = 1
        for row_cells in ws.iter_rows(min_row=2):
            row_cells[item_no_col_idx].value = global_item_no_counter
            global_item_no_counter += 1
            row_cells[section_col_idx].value = ws.title
            row_cells[section_s_no_col_idx].value = section_s_no_counter
            section_s_no_counter += 1
    wb.save(EXCEL_FILE)

def load_logs():
    ensure_log_file()
    wb_log = openpyxl.load_workbook(LOG_FILE)
    rows = list(wb_log.active.values)
    return rows[0], rows[1:]

# --- Routes ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        user_hash = USERS.get(username)
        if user_hash and check_password_hash(user_hash, password):
            session['username'] = username.capitalize()
            session['role'] = 'admin' if username in ADMIN_USERS else 'user'
            return redirect(url_for('index'))
        return render_template('login.html', error='Invalid username or password.')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    headers, data = headers_and_rows()
    all_sections = sorted([ws.title for ws in load_workbook().worksheets])
    try:
        item_no_col_idx = headers.index('Item No')
        sorted_data = sorted(data, key=lambda row: int(row[item_no_col_idx] or 0))
    except (ValueError, IndexError):
        sorted_data = data
    return render_template('index.html', headers=headers, data=sorted_data, session=session, all_sections=all_sections)

@app.route('/edit', methods=['GET'])
@app.route('/edit/<int:item_no>', methods=['GET'])
@login_required
@admin_required
def edit_page(item_no=None):
    headers, data = headers_and_rows()
    all_sections = sorted([ws.title for ws in load_workbook().worksheets])
    item_data = None
    if item_no:
        for row in data:
            if str(row[0]) == str(item_no):
                item_data = {headers[i]: val for i, val in enumerate(row)}
                break
    return render_template('edit.html', item_data=item_data, all_sections=all_sections, session=session)

@app.route('/transaction')
@login_required
def transaction_page():
    headers, data = headers_and_rows()
    try:
        item_no_col_idx = headers.index('Item No')
        sorted_data = sorted(data, key=lambda row: int(row[item_no_col_idx] or 0))
    except (ValueError, IndexError):
        sorted_data = data
    return render_template('transaction.html', headers=headers, data=sorted_data, session=session)

@app.route('/add', methods=['POST'])
@login_required
@admin_required
def add_item():
    payload = request.form
    section = payload.get('Section') or 'Uncategorized'
    
    with LOCK:
        wb = load_workbook()
        headers = [cell.value for cell in wb.worksheets[0][1]] if wb.worksheets else ['Item No','Section','Section S.No','Component','Box No','Specifications','Existing Units','Part number','SKU ID','Remarks','URLS']
        
        if section not in wb.sheetnames:
            ws = wb.create_sheet(title=section)
            ws.append(headers)
        else:
            ws = wb[section]
            
        new_row_values = [None] * len(headers)
        header_map = {h.lower(): i for i, h in enumerate(headers)}

        def set_val(col, val):
            if col.lower() in header_map:
                new_row_values[header_map[col.lower()]] = val

        # Set all string-based values first
        set_val('Section', section)
        set_val('Component', payload.get('Component'))
        set_val('Box No', payload.get('Box No'))
        set_val('Specifications', payload.get('Specifications'))
        set_val('Part number', payload.get('Part number'))
        set_val('SKU ID', payload.get('SKU ID'))
        set_val('URLS', payload.get('URLS'))
        
        # Get the 'Existing Units' value from the form
        units_input = payload.get('Existing Units', '0')
        
        # --- THIS IS THE CORRECTED LOGIC ---
        try:
            # Try to convert to a number for comparison
            units_num = int(float(units_input))
            # If successful, save the NUMBER to the sheet
            set_val('Existing Units', units_num)
            
            # Set remarks based on the NUMBER
            if units_num <= 0:
                remark = 'Out of Stock'
            elif units_num <= 5:
                remark = 'Low Stock'
            else:
                remark = 'In Stock'
            set_val('Remarks', remark)
        except (ValueError, TypeError):
            # If it fails (e.g., input was "In Stock"), save the STRING to the sheet
            set_val('Existing Units', units_input)
            # Set a safe default remark
            set_val('Remarks', 'In Stock')

        ws.append(new_row_values)
        wb.save(EXCEL_FILE)
        save_workbook_with_reindex(load_workbook())

    return redirect(url_for('index'))

@app.route('/delete', methods=['POST'])
@login_required
@admin_required
def delete():
    payload = request.get_json()
    if not payload or 'item_no' not in payload:
        return jsonify({'ok': False, 'error': 'Item number is required.'}), 400
    
    item_no = int(payload.get('item_no'))
    
    with LOCK:
        wb = load_workbook()
        if not wb.sheetnames:
            return jsonify({'ok': False, 'error': 'Inventory is empty.'}), 500

        headers = [cell.value for cell in wb.worksheets[0][1]]
        item_no_col_idx = headers.index('Item No')
        
        row_to_delete_num = None
        sheet_to_delete_from = None

        # Find the row to delete across all sheets
        for sheet in wb.worksheets:
            # Iterate backwards when deleting to avoid index shifting issues
            for r_idx in range(sheet.max_row, 1, -1):
                cell_val = sheet.cell(row=r_idx, column=item_no_col_idx + 1).value
                if str(cell_val) == str(item_no):
                    row_to_delete_num = r_idx
                    sheet_to_delete_from = sheet
                    break
            if row_to_delete_num:
                break
        
        if not row_to_delete_num:
            return jsonify({'ok': False, 'error': 'Item not found.'}), 404
        
        # Delete the identified row
        sheet_to_delete_from.delete_rows(row_to_delete_num)
        
        # If a sheet becomes empty (only has a header), remove it
        if sheet_to_delete_from.max_row < 2 and len(wb.sheetnames) > 1:
            wb.remove(sheet_to_delete_from)

        # Save the changes before re-indexing
        wb.save(EXCEL_FILE)
        
        # Re-index the entire workbook
        save_workbook_with_reindex(load_workbook())

    return jsonify({'ok': True})

@app.route('/update', methods=['POST'])
@login_required
@admin_required
def update():
    payload = request.get_json()
    item_no = int(payload.get('item_no'))
    col_name = payload.get('col_name')
    new_value = payload.get('new_value')
    
    with LOCK:
        wb = load_workbook()
        if not wb.sheetnames: return jsonify({'ok': False, 'error': 'No sheets in workbook'}), 500
        
        headers = [cell.value for cell in wb.worksheets[0][1]]
        item_no_col_idx = headers.index('Item No')
        
        target_row_obj = None
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2):
                if str(row[item_no_col_idx].value) == str(item_no):
                    target_row_obj = row
                    break
            if target_row_obj: break

        if not target_row_obj:
            return jsonify({'ok': False, 'error': 'Item not found'}), 404

        # --- THIS IS THE CORRECTED LOGIC ---
        # Normalize both the list of headers and the incoming column name
        normalized_headers = [str(h).strip().lower() for h in headers]
        normalized_col_name = str(col_name).strip().lower()

        try:
            # Find the index in the normalized list
            col_idx = normalized_headers.index(normalized_col_name)
        except ValueError:
            # If the column name from the frontend doesn't exist in the Excel file
            return jsonify({'ok': False, 'error': f"Column '{col_name}' not found"}), 400

        # Update the cell in the target row using the correct index
        target_row_obj[col_idx].value = new_value

        # The rest of the logic for updating 'Remarks' is now safe
        if normalized_col_name == 'existing units':
            try:
                units = int(float(new_value))
                remark = 'Out of Stock' if units <= 0 else ('Low Stock' if units <= 5 else 'In Stock')
                remarks_col_idx = normalized_headers.index('remarks')
                target_row_obj[remarks_col_idx].value = remark
            except (ValueError, IndexError):
                pass # Fail silently if remarks can't be updated
                
        wb.save(EXCEL_FILE)
        
    return jsonify({'ok': True})

@app.route('/edit_item/<int:item_no>', methods=['POST'])
@login_required
@admin_required
def update_item(item_no):
    payload = request.form
    with LOCK:
        wb = load_workbook()
        headers = [cell.value for cell in wb.worksheets[0][1]]
        item_no_col_idx = headers.index('Item No')
        target_row_obj = None

        # Find the row to edit across all sheets
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2):
                if str(row[item_no_col_idx].value) == str(item_no):
                    target_row_obj = row
                    break
            if target_row_obj: break
        
        if not target_row_obj:
            # Handle error: item not found
            return redirect(url_for('index'))

        # Update values in the found row
        for i, header in enumerate(headers):
            if header in payload:
                # Handle numeric conversion for 'Existing Units'
                if header.lower() == 'existing units':
                    try:
                        target_row_obj[i].value = int(payload[header])
                    except (ValueError, TypeError):
                        target_row_obj[i].value = 0 # Default to 0 if input is invalid
                else:
                    target_row_obj[i].value = payload[header]

        # Recalculate 'Remarks' based on new stock
        units_col_idx = headers.index('Existing Units')
        remarks_col_idx = headers.index('Remarks')
        new_stock = int(target_row_obj[units_col_idx].value or 0)
        target_row_obj[remarks_col_idx].value = 'Out of Stock' if new_stock <= 0 else ('Low Stock' if new_stock <= 5 else 'In Stock')
        
        # This is a complex operation: if the section was changed, we need to move the row
        # For simplicity, we will first re-save and then handle re-indexing which also corrects the Section column
        wb.save(EXCEL_FILE)
        save_workbook_with_reindex(load_workbook()) # Re-run re-indexing
        
    return redirect(url_for('index'))

def perform_transaction(item_no, quantity_change, transaction_type):
    username = session.get('username')
    with LOCK:
        wb_inv = load_workbook()
        headers = [cell.value for cell in wb_inv.worksheets[0][1]]
        item_no_col_idx = headers.index('Item No')
        units_col_idx = headers.index('Existing Units')
        
        target_row_obj = None
        for sheet in wb_inv.worksheets:
            for row in sheet.iter_rows(min_row=2):
                if str(row[item_no_col_idx].value) == str(item_no):
                    target_row_obj = row; break
            if target_row_obj: break

        if not target_row_obj:
            return jsonify({'success': False, 'error': 'Item not found in inventory.'}), 404
        
        try:
            current_stock = int(float(target_row_obj[units_col_idx].value or 0))
        except (ValueError, TypeError):
            # If stock is a string like "In Stock", we cannot perform a transaction
            return jsonify({'success': False, 'error': 'Cannot transact with non-numeric stock items.'}), 400

        if transaction_type == "RETRIEVE" and quantity_change > current_stock:
            return jsonify({'success': False, 'error': f'Not enough stock. Only {current_stock} available.'}), 400

        new_stock = current_stock + quantity_change # quantity_change will be negative for retrievals
        target_row_obj[units_col_idx].value = new_stock
        target_row_obj[headers.index('Remarks')].value = 'Out of Stock' if new_stock <= 0 else ('Low Stock' if new_stock <= 5 else 'In Stock')
        wb_inv.save(EXCEL_FILE)

        ensure_log_file()
        wb_log = openpyxl.load_workbook(LOG_FILE)
        ws_log = wb_log.active
        ws_log.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"), username, transaction_type, item_no,
            target_row_obj[headers.index('Component')].value,
            f"{'+' if quantity_change > 0 else ''}{quantity_change}", new_stock
        ])
        wb_log.save(LOG_FILE)

    return jsonify({'success': True, 'message': f'Transaction successful.'})

@app.route('/retrieve/<int:item_no>', methods=['POST'])
@login_required
def retrieve_item(item_no):
    quantity = request.get_json().get('quantity')
    if not quantity or quantity <= 0: return jsonify({'success': False, 'error': 'Invalid quantity.'}), 400
    return perform_transaction(item_no, -quantity, "RETRIEVE")

@app.route('/return_item/<int:item_no>', methods=['POST'])
@login_required
def return_item(item_no):
    quantity = request.get_json().get('quantity')
    if not quantity or quantity <= 0: return jsonify({'success': False, 'error': 'Invalid quantity.'}), 400
    return perform_transaction(item_no, quantity, "RETURN")

@app.route('/logs')
@login_required
@admin_required
def view_logs():
    log_headers, log_data = load_logs()
    log_data.reverse()
    return render_template('logs.html', headers=log_headers, data=log_data, session=session)

@app.route('/download')
@login_required
@admin_required
def download():
    return send_from_directory(directory=str(BASE_DIR), path=EXCEL_FILE.name, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=6969, debug=True)
# --- END OF FILE ---